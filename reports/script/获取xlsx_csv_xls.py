from openpyxl import load_workbook

filename1 = r'E:\just language\reports\data\a.xlsx'
filename2 = r'E:\just language\reports\data\aa.xls'

# 返回的是一个对象
wb = load_workbook(filename1)
ws = wb.get_sheet_by_name("Sheet1")

print(ws1.cell(row=1,column=1).value)
    for sheet_name in wb.get_named_ranges():
        sheet=wb.get_sheet_by_name(sheet_name)
        for i in range(1,sheet.max_row+1):
            print(sheet.cell(row=i,column=1).value)

for i in range(1, 10):
    print(ws1.cell(row = i, column = 1).value)


# xls文件的处理方式
import xlrd

book = xlrd.open_workbook(filename2)
sheet0 = book.sheet_by_index(0)
sheet_name = book.sheet_names()[0]
print(sheet_name)

sheet1 = book.sheet_by_name(sheet_name)

nrows = sheet0.nrows
ncols = sheet0.ncols

row_value = sheet0.row_value(0)
col_value = sheet0.col_value(0)

cell_value1 = sheet0.cell_value(0, 0)
cell_value2 = sheet0.cell_value(0, 1)
